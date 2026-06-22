"""ExcelReportSink — ReportSink adapter that writes openpyxl workbooks.

Produces one ``.xlsx`` file per trading day with four sheets:
  - Summary:   NAV, P&L, benchmark return, alpha, trade count
  - Positions: symbol, qty, avg_cost, current_price, unrealized_pnl
  - Trades:    per-trade execution details
  - Evidence:  citation metadata (source_url, chunk_id, published_at, symbol)

``send_hitl_request`` is intentionally not supported — use SlackReportSink for
interactive HITL.  ``send_alert`` appends a row to a shared CSV in the output
directory so operational alerts survive without a Slack connection.
"""

from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from firm.ports.types import ApprovalResult, DailyReport, HITLRequest

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_ALERTS_FILENAME = "alerts.csv"
_ALERTS_HEADER = ["ts", "correlation_id", "message"]


# ---------------------------------------------------------------------------
# Sheet writers — one function per sheet; each ≤30 lines
# ---------------------------------------------------------------------------


def _write_summary_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Populate the Summary sheet with one-row daily KPIs."""
    ws.title = "Summary"
    headers = ["date", "nav", "pnl", "benchmark_return", "alpha", "num_trades"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    alpha = float(report.pnl / report.nav) - report.benchmark_return if report.nav else 0.0
    ws.append(
        [
            str(report.date),
            float(report.nav),
            float(report.pnl),
            report.benchmark_return,
            alpha,
            len(report.trades),
        ]
    )


def _write_positions_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Populate the Positions sheet, one row per holding."""
    ws.title = "Positions"
    headers = ["symbol", "qty", "avg_cost", "current_price", "unrealized_pnl"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for pos in report.positions:
        qty = float(pos.get("qty", 0))
        avg_cost = float(pos.get("avg_cost", 0))
        current_price = float(pos.get("current_price", 0))
        # Prefer the upstream-computed value; fall back to local computation only
        # when absent (e.g., in legacy fixtures that pre-date the field).
        unrealized_pnl = float(
            pos["unrealized_pnl"] if "unrealized_pnl" in pos else (current_price - avg_cost) * qty
        )
        ws.append(
            [
                pos.get("symbol", ""),
                qty,
                avg_cost,
                current_price,
                unrealized_pnl,
            ]
        )


def _write_trades_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Populate the Trades sheet, one row per execution."""
    ws.title = "Trades"
    headers = [
        "cycle_id",
        "symbol",
        "side",
        "qty",
        "fill_price",
        "slippage",
        "commission",
        "status",
        "ts",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for trade in report.trades:
        ws.append(
            [
                str(trade.get("cycle_id", "")),
                trade.get("symbol", ""),
                trade.get("side", ""),
                float(trade.get("qty", 0)),
                float(trade.get("fill_price", 0)),
                float(trade.get("slippage", 0)),
                float(trade.get("commission", 0)),
                trade.get("status", ""),
                str(trade.get("ts", "")),
            ]
        )


def _write_evidence_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Populate the Evidence sheet with citation metadata."""
    ws.title = "Evidence"
    headers = ["source_url", "chunk_id", "published_at", "symbol"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for citation in report.citations:
        ws.append(
            [
                citation.get("source_url", ""),
                citation.get("chunk_id", ""),
                str(citation.get("published_at", "")),
                citation.get("symbol", ""),
            ]
        )


# ---------------------------------------------------------------------------
# ExcelReportSink
# ---------------------------------------------------------------------------


class ExcelReportSink:
    """Implements :class:`~firm.ports.report.ReportSink` via openpyxl.

    Each call to ``send_daily_report`` produces one ``YYYY-MM-DD.xlsx`` file
    in *output_dir*.  ``send_alert`` appends to ``alerts.csv`` in the same
    directory so alerts are durable without a network connection.
    """

    def __init__(self, output_dir: Path | None = None) -> None:
        self._output_dir = output_dir if output_dir is not None else Path("data/reports").resolve()
        self._output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # ReportSink protocol
    # ------------------------------------------------------------------

    def send_daily_report(self, report: DailyReport) -> None:
        """Write *report* to ``<output_dir>/<report.date>.xlsx``."""
        wb = openpyxl.Workbook()
        # openpyxl creates a default "Sheet" — use it for Summary, add the rest.
        assert wb.active is not None
        summary_ws: Worksheet = wb.active
        _write_summary_sheet(summary_ws, report)

        _write_positions_sheet(wb.create_sheet(), report)
        _write_trades_sheet(wb.create_sheet(), report)
        _write_evidence_sheet(wb.create_sheet(), report)

        filename = self._output_dir / f"{report.date}.xlsx"
        wb.save(filename)

    def send_hitl_request(self, req: HITLRequest) -> ApprovalResult:
        """Not supported — Excel cannot receive interactive approvals."""
        raise NotImplementedError("Excel sink does not support HITL — use SlackReportSink")

    def send_alert(self, message: str, correlation_id: str) -> None:
        """Append one row to ``<output_dir>/alerts.csv``."""
        alerts_path = self._output_dir / _ALERTS_FILENAME
        is_new_file = not alerts_path.exists()
        with alerts_path.open("a", newline="") as fh:
            writer = csv.writer(fh)
            if is_new_file:
                writer.writerow(_ALERTS_HEADER)
            writer.writerow(
                [
                    datetime.now(tz=UTC).isoformat(),
                    correlation_id,
                    message,
                ]
            )
