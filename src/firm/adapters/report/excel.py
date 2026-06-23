from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from firm.ports.types import ApprovalResult, DailyReport, HITLRequest

_ALERTS_FILENAME = "alerts.csv"
_ALERTS_HEADER = ["ts", "correlation_id", "message"]


class ExcelReportSink:
    def __init__(self, output_dir: Path | None = None) -> None:
        self._output_dir = output_dir if output_dir is not None else Path("data/reports").resolve()
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def send_daily_report(self, report: DailyReport) -> None:
        wb = openpyxl.Workbook()
        assert wb.active is not None
        _write_summary(wb.active, report)
        _write_positions(wb.create_sheet(), report)
        _write_trades(wb.create_sheet(), report)
        _write_evidence(wb.create_sheet(), report)
        wb.save(self._output_dir / f"{report.date}.xlsx")

    def send_hitl_request(self, req: HITLRequest) -> ApprovalResult:
        raise NotImplementedError("Excel sink does not support HITL — use SlackReportSink")

    def send_alert(self, message: str, correlation_id: str) -> None:
        alerts_path = self._output_dir / _ALERTS_FILENAME
        is_new = not alerts_path.exists()
        with alerts_path.open("a", newline="") as fh:
            writer = csv.writer(fh)
            if is_new:
                writer.writerow(_ALERTS_HEADER)
            writer.writerow([datetime.now(tz=UTC).isoformat(), correlation_id, message])


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _bold_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_summary(ws: Worksheet, report: DailyReport) -> None:
    ws.title = "Summary"
    _bold_header(ws, ["date", "nav", "pnl", "benchmark_return", "alpha", "num_trades"])
    alpha = float(report.pnl / report.nav) - report.benchmark_return if report.nav else 0.0
    ws.append(
        [
            str(report.date),
            float(report.nav),
            float(report.pnl),
            report.benchmark_return,
            alpha,
            len(report.trades),
        ]
    )


def _write_positions(ws: Worksheet, report: DailyReport) -> None:
    ws.title = "Positions"
    _bold_header(ws, ["symbol", "qty", "avg_cost", "current_price", "unrealized_pnl"])
    for p in report.positions:
        ws.append([p["symbol"], p["qty"], p["avg_cost"], p["current_price"], p["unrealized_pnl"]])


def _write_trades(ws: Worksheet, report: DailyReport) -> None:
    ws.title = "Trades"
    _bold_header(
        ws,
        [
            "cycle_id",
            "symbol",
            "side",
            "qty",
            "fill_price",
            "slippage",
            "commission",
            "status",
            "ts",
        ],
    )
    for t in report.trades:
        ws.append(
            [
                t["cycle_id"],
                t["symbol"],
                t["side"],
                t["qty"],
                t["fill_price"],
                t["slippage"],
                t["commission"],
                t["status"],
                t.get("ts", ""),
            ]
        )


def _write_evidence(ws: Worksheet, report: DailyReport) -> None:
    ws.title = "Evidence"
    _bold_header(ws, ["source_url", "chunk_id", "published_at", "symbol"])
    for c in report.citations:
        ws.append([c["source_url"], c["chunk_id"], c["published_at"], c["symbol"]])
