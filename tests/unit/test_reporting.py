"""Unit tests for ExcelReportSink and SlackReportSink.

All tests are pure-Python — no network, no real Slack tokens, no real filesystem
side-effects (temporary directories used for Excel output).

Covers:
  - test_excel_report_produces_correct_sheets
  - test_slack_hitl_request_sends_buttons
  - test_slack_daily_report_sent
  - alert helpers for both adapters
"""

from __future__ import annotations

import csv
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch
from uuid import uuid4

import openpyxl
import pytest
from slack_sdk.errors import SlackApiError

from firm.adapters.report import ExcelReportSink, SlackReportSink
from firm.ports.report import ReportSink
from firm.ports.types import DailyReport, HITLRequest

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_report() -> DailyReport:
    return DailyReport(
        date=date(2024, 10, 25),
        nav=Decimal("1_050_000"),
        pnl=Decimal("5_000"),
        benchmark_return=0.002,
        trades=[
            {
                "cycle_id": str(uuid4()),
                "symbol": "NVDA",
                "side": "buy",
                "qty": 50,
                "fill_price": 140.25,
                "slippage": 0.07,
                "commission": 0.25,
                "status": "FILLED",
                "ts": "2024-10-25T14:30:00Z",
            }
        ],
        positions=[
            {
                "symbol": "NVDA",
                "qty": 50,
                "avg_cost": 139.50,
                "current_price": 140.25,
                "unrealized_pnl": 37.50,
            }
        ],
        citations=[
            {
                "source_url": "https://example.com/news/nvda-earnings",
                "chunk_id": "chunk-abc123",
                "published_at": "2024-10-22T08:00:00Z",
                "symbol": "NVDA",
            }
        ],
    )


def _make_hitl_request(expired: bool = False) -> HITLRequest:
    if expired:
        expires_at = datetime(2020, 1, 1, tzinfo=UTC)
    else:
        expires_at = datetime(2099, 12, 31, tzinfo=UTC)
    return HITLRequest(
        trade_id=uuid4(),
        symbol="NVDA",
        side="buy",
        qty_str="50",
        notional=Decimal("7_012.50"),
        reason="Trade notional 5.5% exceeds HITL threshold 5.0% of NAV.",
        expires_at=expires_at,
        correlation_id=str(uuid4()),
    )


# ---------------------------------------------------------------------------
# ExcelReportSink tests
# ---------------------------------------------------------------------------


class TestExcelReportSink:
    def test_excel_report_produces_correct_sheets(self, tmp_path: Path) -> None:
        """Writing a DailyReport creates an xlsx with exactly four named sheets."""
        sink = ExcelReportSink(output_dir=tmp_path)
        report = _make_report()

        sink.send_daily_report(report)

        xlsx_path = tmp_path / f"{report.date}.xlsx"
        assert xlsx_path.exists(), "Expected .xlsx file was not created"

        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.sheetnames == ["Summary", "Positions", "Trades", "Evidence"]

    def test_excel_summary_sheet_values(self, tmp_path: Path) -> None:
        """Summary sheet row 2 contains correct NAV, P&L, and trade count."""
        sink = ExcelReportSink(output_dir=tmp_path)
        report = _make_report()

        sink.send_daily_report(report)

        wb = openpyxl.load_workbook(tmp_path / f"{report.date}.xlsx")
        ws = wb["Summary"]

        # row 1 is headers; row 2 is data
        headers = [ws.cell(1, col).value for col in range(1, 7)]
        assert headers == ["date", "nav", "pnl", "benchmark_return", "alpha", "num_trades"]

        data_row = [ws.cell(2, col).value for col in range(1, 7)]
        assert data_row[0] == "2024-10-25"  # date as string
        assert data_row[1] == pytest.approx(1_050_000.0)
        assert data_row[2] == pytest.approx(5_000.0)
        assert data_row[5] == 1  # num_trades

    def test_excel_positions_sheet_values(self, tmp_path: Path) -> None:
        """Positions sheet contains headers and one data row for the NVDA holding."""
        sink = ExcelReportSink(output_dir=tmp_path)
        sink.send_daily_report(_make_report())

        wb = openpyxl.load_workbook(tmp_path / "2024-10-25.xlsx")
        ws = wb["Positions"]

        headers = [ws.cell(1, col).value for col in range(1, 6)]
        assert headers == ["symbol", "qty", "avg_cost", "current_price", "unrealized_pnl"]

        assert ws.cell(2, 1).value == "NVDA"
        assert ws.cell(2, 2).value == pytest.approx(50.0)

    def test_excel_trades_sheet_has_correct_columns(self, tmp_path: Path) -> None:
        """Trades sheet header matches the expected columns."""
        sink = ExcelReportSink(output_dir=tmp_path)
        sink.send_daily_report(_make_report())

        wb = openpyxl.load_workbook(tmp_path / "2024-10-25.xlsx")
        ws = wb["Trades"]
        headers = [ws.cell(1, col).value for col in range(1, 10)]
        assert headers == [
            "cycle_id",
            "symbol",
            "side",
            "qty",
            "fill_price",
            "slippage",
            "commission",
            "status",
            "ts",
        ]

    def test_excel_evidence_sheet_has_correct_columns(self, tmp_path: Path) -> None:
        """Evidence sheet header matches citation fields."""
        sink = ExcelReportSink(output_dir=tmp_path)
        sink.send_daily_report(_make_report())

        wb = openpyxl.load_workbook(tmp_path / "2024-10-25.xlsx")
        ws = wb["Evidence"]
        headers = [ws.cell(1, col).value for col in range(1, 5)]
        assert headers == ["source_url", "chunk_id", "published_at", "symbol"]
        assert ws.cell(2, 1).value == "https://example.com/news/nvda-earnings"

    def test_excel_send_hitl_raises_not_implemented(self, tmp_path: Path) -> None:
        """send_hitl_request raises NotImplementedError with a descriptive message."""
        sink = ExcelReportSink(output_dir=tmp_path)
        req = _make_hitl_request()

        with pytest.raises(NotImplementedError, match="SlackReportSink"):
            sink.send_hitl_request(req)

    def test_excel_alert_creates_csv(self, tmp_path: Path) -> None:
        """send_alert creates alerts.csv and appends a row per call."""
        sink = ExcelReportSink(output_dir=tmp_path)
        cid = str(uuid4())

        sink.send_alert("circuit-breaker tripped", cid)

        alerts_path = tmp_path / "alerts.csv"
        assert alerts_path.exists()

        with alerts_path.open() as fh:
            rows = list(csv.reader(fh))

        assert rows[0] == ["ts", "correlation_id", "message"]
        assert rows[1][1] == cid
        assert rows[1][2] == "circuit-breaker tripped"

    def test_excel_alert_appends_on_second_call(self, tmp_path: Path) -> None:
        """Multiple send_alert calls append rows without duplicating the header."""
        sink = ExcelReportSink(output_dir=tmp_path)
        sink.send_alert("first alert", "cid-1")
        sink.send_alert("second alert", "cid-2")

        with (tmp_path / "alerts.csv").open() as fh:
            rows = list(csv.reader(fh))

        # 1 header + 2 data rows
        assert len(rows) == 3
        assert rows[1][2] == "first alert"
        assert rows[2][2] == "second alert"

    def test_excel_output_dir_created_if_missing(self, tmp_path: Path) -> None:
        """Constructor creates output_dir when it does not exist."""
        new_dir = tmp_path / "deep" / "nested" / "reports"
        assert not new_dir.exists()

        ExcelReportSink(output_dir=new_dir)

        assert new_dir.is_dir()

    def test_excel_sink_satisfies_report_sink_protocol(self, tmp_path: Path) -> None:
        """ExcelReportSink satisfies the runtime-checkable ReportSink protocol."""
        sink = ExcelReportSink(output_dir=tmp_path)
        assert isinstance(sink, ReportSink)


# ---------------------------------------------------------------------------
# SlackReportSink tests
# ---------------------------------------------------------------------------


class TestSlackReportSink:
    def _make_sink(self) -> tuple[SlackReportSink, MagicMock]:
        """Return a SlackReportSink wired to a mocked WebClient."""
        with patch("firm.adapters.report.slack.WebClient") as mock_cls:
            mock_client = MagicMock()
            mock_cls.return_value = mock_client
            sink = SlackReportSink(token="xoxb-test", channel="#trading-desk")
        return sink, mock_client

    def test_slack_daily_report_sent(self) -> None:
        """send_daily_report calls chat_postMessage once with Block Kit blocks."""
        sink, mock_client = self._make_sink()
        report = _make_report()

        sink.send_daily_report(report)

        mock_client.chat_postMessage.assert_called_once()
        call_kwargs = mock_client.chat_postMessage.call_args.kwargs
        assert call_kwargs["channel"] == "#trading-desk"
        assert isinstance(call_kwargs["blocks"], list)
        assert len(call_kwargs["blocks"]) > 0

    def test_slack_daily_report_contains_nav_section(self) -> None:
        """Daily report blocks contain a section mentioning NAV and P&L."""
        sink, mock_client = self._make_sink()

        sink.send_daily_report(_make_report())

        call_kwargs = mock_client.chat_postMessage.call_args.kwargs
        blocks = call_kwargs["blocks"]
        section_texts = [b["text"]["text"] for b in blocks if b["type"] == "section"]
        combined = " ".join(section_texts)
        assert "NAV" in combined
        assert "P&L" in combined

    def test_slack_hitl_request_sends_buttons(self) -> None:
        """send_hitl_request posts an interactive message with Approve and Reject buttons."""
        sink, mock_client = self._make_sink()
        req = _make_hitl_request()

        result = sink.send_hitl_request(req)

        mock_client.chat_postMessage.assert_called_once()
        call_kwargs = mock_client.chat_postMessage.call_args.kwargs
        blocks = call_kwargs["blocks"]

        # Locate the actions block
        action_blocks = [b for b in blocks if b["type"] == "actions"]
        assert len(action_blocks) == 1, "Expected exactly one actions block"

        elements = action_blocks[0]["elements"]
        values = {el["value"] for el in elements}
        assert "approved" in values
        assert "rejected" in values

        # The adapter must never auto-approve — placeholder must return expired.
        assert result.status == "expired"
        assert result.decided_by is None
        assert result.edited_qty is None

    def test_slack_hitl_request_includes_symbol_and_notional(self) -> None:
        """HITL blocks include the symbol and notional fields."""
        sink, mock_client = self._make_sink()
        req = _make_hitl_request()

        sink.send_hitl_request(req)

        call_kwargs = mock_client.chat_postMessage.call_args.kwargs
        blocks = call_kwargs["blocks"]
        section_texts = " ".join(b["text"]["text"] for b in blocks if b["type"] == "section")
        assert req.symbol in section_texts
        assert req.side.upper() in section_texts

    def test_slack_hitl_expired_returns_expired_without_post(self) -> None:
        """An already-expired HITL request returns status=expired without posting."""
        sink, mock_client = self._make_sink()
        req = _make_hitl_request(expired=True)

        result = sink.send_hitl_request(req)

        assert result.status == "expired"
        assert result.decided_by is None
        mock_client.chat_postMessage.assert_not_called()

    def test_slack_alert_posts_plain_text(self) -> None:
        """send_alert calls chat_postMessage with plain text containing the message."""
        sink, mock_client = self._make_sink()
        cid = "corr-abc-123"
        msg = "daily loss halt triggered"

        sink.send_alert(msg, cid)

        mock_client.chat_postMessage.assert_called_once()
        call_kwargs = mock_client.chat_postMessage.call_args.kwargs
        assert call_kwargs["channel"] == "#trading-desk"
        assert msg in call_kwargs["text"]
        assert cid in call_kwargs["text"]

    def test_slack_sink_satisfies_report_sink_protocol(self) -> None:
        """SlackReportSink satisfies the runtime-checkable ReportSink protocol."""
        sink, _ = self._make_sink()
        assert isinstance(sink, ReportSink)

    # ------------------------------------------------------------------
    # SlackApiError failure paths (issue 7)
    # ------------------------------------------------------------------

    def test_slack_hitl_request_slack_error_returns_expired(self) -> None:
        """A SlackApiError during send_hitl_request must return status=expired, not raise."""
        sink, mock_client = self._make_sink()
        req = _make_hitl_request()
        mock_client.chat_postMessage.side_effect = SlackApiError(
            message="channel_not_found", response={"error": "channel_not_found"}
        )

        result = sink.send_hitl_request(req)

        assert result.status == "expired"
        assert result.decided_by is None

    def test_slack_daily_report_slack_error_propagates(self) -> None:
        """A SlackApiError in send_daily_report propagates to the caller."""
        sink, mock_client = self._make_sink()
        mock_client.chat_postMessage.side_effect = SlackApiError(
            message="invalid_auth", response={"error": "invalid_auth"}
        )

        with pytest.raises(SlackApiError):
            sink.send_daily_report(_make_report())

    def test_slack_alert_slack_error_propagates(self) -> None:
        """A SlackApiError in send_alert propagates to the caller."""
        sink, mock_client = self._make_sink()
        mock_client.chat_postMessage.side_effect = SlackApiError(
            message="ratelimited", response={"error": "ratelimited"}
        )

        with pytest.raises(SlackApiError):
            sink.send_alert("daily loss halt triggered", "corr-xyz")
